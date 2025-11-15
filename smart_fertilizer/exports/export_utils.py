import json
import csv
import xml.etree.ElementTree as ET
from typing import Dict, List, Optional, Union
from datetime import datetime
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd

class ExportUtilities:
    """
    Utility class for exporting fertilizer recommendations in various formats
    """
    
    def __init__(self):
        self.supported_formats = ['json', 'csv', 'xlsx', 'xml', 'txt', 'pdf']
        
    def export_recommendation(self, recommendation: Dict, format_type: str, 
                            include_metadata: bool = True) -> Union[str, bytes]:
        """
        Export recommendation in specified format
        """
        
        format_type = format_type.lower()
        
        if format_type not in self.supported_formats:
            raise ValueError(f"Unsupported format: {format_type}")
        
        # Prepare data
        export_data = self._prepare_export_data(recommendation, include_metadata)
        
        if format_type == 'json':
            return self._export_json(export_data)
        elif format_type == 'csv':
            return self._export_csv(export_data)
        elif format_type == 'xlsx':
            return self._export_xlsx(export_data)
        elif format_type == 'xml':
            return self._export_xml(export_data)
        elif format_type == 'txt':
            return self._export_txt(export_data)
        elif format_type == 'pdf':
            from .pdf_generator import FertilizerReportGenerator
            generator = FertilizerReportGenerator()
            return generator.generate_fertilizer_report(recommendation)
        else:
            raise ValueError(f"Export method not implemented for: {format_type}")
    
    def _prepare_export_data(self, recommendation: Dict, include_metadata: bool) -> Dict:
        """Prepare data for export"""
        
        export_data = {
            'basic_info': {
                'recommendation_id': recommendation.get('recommendation_id', ''),
                'generated_at': str(recommendation.get('generated_at', datetime.now())),
                'region': recommendation.get('region', ''),
                'area_hectares': recommendation.get('area_hectares', 0),
                'target_yield': recommendation.get('target_yield', 0),
                'expected_yield': recommendation.get('expected_yield', 0),
                'roi_percentage': recommendation.get('roi_percentage', 0)
            }
        }
        
        # Soil analysis
        soil_analysis = recommendation.get('soil_analysis', {})
        export_data['soil_analysis'] = {
            'ph': getattr(soil_analysis, 'ph', 0),
            'organic_matter': getattr(soil_analysis, 'organic_matter', 0),
            'nitrogen': getattr(soil_analysis, 'nitrogen', 0),
            'phosphorus': getattr(soil_analysis, 'phosphorus', 0),
            'potassium': getattr(soil_analysis, 'potassium', 0),
            'cec': getattr(soil_analysis, 'cec', 0),
            'texture': getattr(soil_analysis, 'texture', ''),
            'ec': getattr(soil_analysis, 'ec', 0),
            'calcium': getattr(soil_analysis, 'calcium', 0),
            'magnesium': getattr(soil_analysis, 'magnesium', 0),
            'sulfur': getattr(soil_analysis, 'sulfur', 0)
        }
        
        # Crop selection
        crop_selection = recommendation.get('crop_selection', {})
        export_data['crop_selection'] = {
            'crop_type': getattr(crop_selection, 'crop_type', ''),
            'variety': getattr(crop_selection, 'variety', ''),
            'planting_season': getattr(crop_selection, 'planting_season', ''),
            'growth_duration': getattr(crop_selection, 'growth_duration', 0)
        }
        
        # Nutrient balance
        nutrient_balance = recommendation.get('nutrient_balance', {})
        export_data['nutrient_requirements'] = {
            'total_nitrogen_kg_per_ha': getattr(nutrient_balance, 'total_n', 0),
            'total_phosphorus_kg_per_ha': getattr(nutrient_balance, 'total_p', 0),
            'total_potassium_kg_per_ha': getattr(nutrient_balance, 'total_k', 0),
            'secondary_nutrients': getattr(nutrient_balance, 'secondary_nutrients', {}),
            'micronutrients': getattr(nutrient_balance, 'micronutrients', {})
        }
        
        # Recommended fertilizers
        recommended_fertilizers = recommendation.get('recommended_fertilizers', [])
        export_data['recommended_fertilizers'] = []
        for fert in recommended_fertilizers:
            export_data['recommended_fertilizers'].append({
                'name': getattr(fert, 'name', ''),
                'n_content': getattr(fert, 'n_content', 0),
                'p_content': getattr(fert, 'p_content', 0),
                'k_content': getattr(fert, 'k_content', 0),
                'price_per_kg': getattr(fert, 'price_per_kg', 0),
                'availability': getattr(fert, 'availability', '')
            })
        
        # Application schedule
        application_schedule = recommendation.get('application_schedule', [])
        export_data['application_schedule'] = []
        for app in application_schedule:
            export_data['application_schedule'].append({
                'stage': getattr(app, 'stage', ''),
                'days_after_planting': getattr(app, 'days_after_planting', 0),
                'fertilizer_type': getattr(app, 'fertilizer_type', ''),
                'amount_kg_per_ha': getattr(app, 'amount_kg_per_ha', 0),
                'application_method': getattr(app, 'application_method', ''),
                'notes': getattr(app, 'notes', '')
            })
        
        # Cost analysis
        cost_analysis = recommendation.get('cost_analysis', {})
        export_data['cost_analysis'] = {
            'total_cost': getattr(cost_analysis, 'total_cost', 0),
            'cost_per_hectare': getattr(cost_analysis, 'cost_per_hectare', 0),
            'currency': getattr(cost_analysis, 'currency', 'USD'),
            'fertilizer_breakdown': getattr(cost_analysis, 'fertilizer_breakdown', {})
        }
        
        # Additional information
        if include_metadata:
            export_data['climate_considerations'] = recommendation.get('climate_considerations', [])
            export_data['risk_factors'] = recommendation.get('risk_factors', [])
            export_data['alternative_options'] = recommendation.get('alternative_options', [])
            export_data['language'] = recommendation.get('language', 'en')
            export_data['units'] = recommendation.get('units', 'metric')
        
        return export_data
    
    def _export_json(self, data: Dict) -> str:
        """Export as JSON"""
        return json.dumps(data, indent=2, ensure_ascii=False, default=str)
    
    def _export_csv(self, data: Dict) -> str:
        """Export as CSV"""
        output = io.StringIO()
        
        # Write basic information
        writer = csv.writer(output)
        writer.writerow(['SMART FERTILIZER RECOMMENDATION REPORT'])
        writer.writerow([''])
        
        # Basic info
        writer.writerow(['BASIC INFORMATION'])
        basic_info = data.get('basic_info', {})
        for key, value in basic_info.items():
            writer.writerow([key.replace('_', ' ').title(), value])
        writer.writerow([''])
        
        # Soil analysis
        writer.writerow(['SOIL ANALYSIS'])
        writer.writerow(['Parameter', 'Value', 'Unit'])
        soil_analysis = data.get('soil_analysis', {})
        
        parameter_units = {
            'ph': '',
            'organic_matter': '%',
            'nitrogen': 'ppm',
            'phosphorus': 'ppm',
            'potassium': 'ppm',
            'cec': 'cmol/kg',
            'texture': '',
            'ec': 'dS/m',
            'calcium': 'ppm',
            'magnesium': 'ppm',
            'sulfur': 'ppm'
        }
        
        for param, value in soil_analysis.items():
            unit = parameter_units.get(param, '')
            writer.writerow([param.replace('_', ' ').title(), value, unit])
        writer.writerow([''])
        
        # Nutrient requirements
        writer.writerow(['NUTRIENT REQUIREMENTS'])
        writer.writerow(['Nutrient', 'Amount (kg/ha)'])
        nutrients = data.get('nutrient_requirements', {})
        writer.writerow(['Nitrogen', nutrients.get('total_nitrogen_kg_per_ha', 0)])
        writer.writerow(['Phosphorus', nutrients.get('total_phosphorus_kg_per_ha', 0)])
        writer.writerow(['Potassium', nutrients.get('total_potassium_kg_per_ha', 0)])
        writer.writerow([''])
        
        # Recommended fertilizers
        writer.writerow(['RECOMMENDED FERTILIZERS'])
        writer.writerow(['Name', 'N%', 'P%', 'K%', 'Price/kg', 'Availability'])
        fertilizers = data.get('recommended_fertilizers', [])
        for fert in fertilizers:
            writer.writerow([
                fert.get('name', ''),
                fert.get('n_content', 0),
                fert.get('p_content', 0),
                fert.get('k_content', 0),
                fert.get('price_per_kg', 0),
                fert.get('availability', '')
            ])
        writer.writerow([''])
        
        # Application schedule
        writer.writerow(['APPLICATION SCHEDULE'])
        writer.writerow(['Stage', 'Days After Planting', 'Fertilizer', 'Rate (kg/ha)', 'Method'])
        schedule = data.get('application_schedule', [])
        for app in schedule:
            writer.writerow([
                app.get('stage', ''),
                app.get('days_after_planting', 0),
                app.get('fertilizer_type', ''),
                app.get('amount_kg_per_ha', 0),
                app.get('application_method', '')
            ])
        writer.writerow([''])
        
        # Cost analysis
        writer.writerow(['COST ANALYSIS'])
        cost_analysis = data.get('cost_analysis', {})
        writer.writerow(['Total Cost', cost_analysis.get('total_cost', 0)])
        writer.writerow(['Cost per Hectare', cost_analysis.get('cost_per_hectare', 0)])
        writer.writerow(['Currency', cost_analysis.get('currency', 'USD')])
        
        return output.getvalue()
    
    def _export_xlsx(self, data: Dict) -> bytes:
        """Export as Excel file"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="2E8B57", end_color="2E8B57", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        self._create_summary_sheet(ws_summary, data, header_font, header_fill, border)
        
        # Soil analysis sheet
        ws_soil = wb.create_sheet("Soil Analysis")
        self._create_soil_sheet(ws_soil, data, header_font, header_fill, border)
        
        # Fertilizer recommendations sheet
        ws_fert = wb.create_sheet("Fertilizers")
        self._create_fertilizer_sheet(ws_fert, data, header_font, header_fill, border)
        
        # Application schedule sheet
        ws_schedule = wb.create_sheet("Application Schedule")
        self._create_schedule_sheet(ws_schedule, data, header_font, header_fill, border)
        
        # Cost analysis sheet
        ws_cost = wb.create_sheet("Cost Analysis")
        self._create_cost_sheet(ws_cost, data, header_font, header_fill, border)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output.getvalue()
    
    def _create_summary_sheet(self, ws, data, header_font, header_fill, border):
        """Create summary sheet"""
        ws.title = "Summary"
        
        # Title
        ws['A1'] = "Smart Fertilizer Recommendation Summary"
        ws['A1'].font = Font(bold=True, size=16)
        ws.merge_cells('A1:C1')
        
        # Basic info
        row = 3
        basic_info = data.get('basic_info', {})
        for key, value in basic_info.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
        
        # Nutrient summary
        row += 2
        ws[f'A{row}'] = "Nutrient Requirements (kg/ha)"
        ws[f'A{row}'].font = header_font
        ws[f'A{row}'].fill = header_fill
        ws.merge_cells(f'A{row}:B{row}')
        
        row += 1
        nutrients = data.get('nutrient_requirements', {})
        ws[f'A{row}'] = "Nitrogen"
        ws[f'B{row}'] = nutrients.get('total_nitrogen_kg_per_ha', 0)
        row += 1
        ws[f'A{row}'] = "Phosphorus"
        ws[f'B{row}'] = nutrients.get('total_phosphorus_kg_per_ha', 0)
        row += 1
        ws[f'A{row}'] = "Potassium"
        ws[f'B{row}'] = nutrients.get('total_potassium_kg_per_ha', 0)
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_soil_sheet(self, ws, data, header_font, header_fill, border):
        """Create soil analysis sheet"""
        ws.title = "Soil Analysis"
        
        # Headers
        headers = ['Parameter', 'Value', 'Unit', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        soil_data = data.get('soil_analysis', {})
        parameter_units = {
            'ph': ('', 'pH'),
            'organic_matter': ('%', 'Organic Matter'),
            'nitrogen': ('ppm', 'Available Nitrogen'),
            'phosphorus': ('ppm', 'Available Phosphorus'),
            'potassium': ('ppm', 'Available Potassium'),
            'cec': ('cmol/kg', 'CEC'),
            'texture': ('', 'Soil Texture'),
            'ec': ('dS/m', 'Electrical Conductivity'),
            'calcium': ('ppm', 'Available Calcium'),
            'magnesium': ('ppm', 'Available Magnesium'),
            'sulfur': ('ppm', 'Available Sulfur')
        }
        
        row = 2
        for param, value in soil_data.items():
            if param in parameter_units:
                unit, display_name = parameter_units[param]
                ws.cell(row=row, column=1, value=display_name).border = border
                ws.cell(row=row, column=2, value=value).border = border
                ws.cell(row=row, column=3, value=unit).border = border
                ws.cell(row=row, column=4, value=self._get_parameter_status(param, value)).border = border
                row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_fertilizer_sheet(self, ws, data, header_font, header_fill, border):
        """Create fertilizer recommendations sheet"""
        ws.title = "Fertilizer Recommendations"
        
        # Headers
        headers = ['Fertilizer Name', 'N Content (%)', 'P Content (%)', 'K Content (%)', 'Price per kg', 'Availability']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        fertilizers = data.get('recommended_fertilizers', [])
        for row, fert in enumerate(fertilizers, 2):
            ws.cell(row=row, column=1, value=fert.get('name', '')).border = border
            ws.cell(row=row, column=2, value=fert.get('n_content', 0)).border = border
            ws.cell(row=row, column=3, value=fert.get('p_content', 0)).border = border
            ws.cell(row=row, column=4, value=fert.get('k_content', 0)).border = border
            ws.cell(row=row, column=5, value=fert.get('price_per_kg', 0)).border = border
            ws.cell(row=row, column=6, value=fert.get('availability', '')).border = border
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 25)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_schedule_sheet(self, ws, data, header_font, header_fill, border):
        """Create application schedule sheet"""
        ws.title = "Application Schedule"
        
        # Headers
        headers = ['Growth Stage', 'Days After Planting', 'Fertilizer Type', 'Rate (kg/ha)', 'Application Method', 'Notes']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        # Data
        schedule = data.get('application_schedule', [])
        for row, app in enumerate(schedule, 2):
            ws.cell(row=row, column=1, value=app.get('stage', '').replace('_', ' ').title()).border = border
            ws.cell(row=row, column=2, value=app.get('days_after_planting', 0)).border = border
            ws.cell(row=row, column=3, value=app.get('fertilizer_type', '')).border = border
            ws.cell(row=row, column=4, value=app.get('amount_kg_per_ha', 0)).border = border
            ws.cell(row=row, column=5, value=app.get('application_method', '').replace('_', ' ').title()).border = border
            ws.cell(row=row, column=6, value=app.get('notes', '')).border = border
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 40)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _create_cost_sheet(self, ws, data, header_font, header_fill, border):
        """Create cost analysis sheet"""
        ws.title = "Cost Analysis"
        
        # Cost summary
        cost_analysis = data.get('cost_analysis', {})
        
        ws['A1'] = "Cost Summary"
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        ws['A3'] = "Total Cost"
        ws['B3'] = cost_analysis.get('total_cost', 0)
        ws['A4'] = "Cost per Hectare"
        ws['B4'] = cost_analysis.get('cost_per_hectare', 0)
        ws['A5'] = "Currency"
        ws['B5'] = cost_analysis.get('currency', 'USD')
        
        # Cost breakdown
        row = 7
        ws[f'A{row}'] = "Cost Breakdown by Fertilizer"
        ws[f'A{row}'].font = Font(bold=True)
        ws.merge_cells(f'A{row}:C{row}')
        
        row += 1
        ws[f'A{row}'] = "Fertilizer Type"
        ws[f'B{row}'] = "Cost"
        ws[f'C{row}'] = "Percentage"
        
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
        
        breakdown = cost_analysis.get('fertilizer_breakdown', {})
        total_cost = cost_analysis.get('total_cost', 1)
        
        row += 1
        for fert_type, cost in breakdown.items():
            percentage = (cost / total_cost) * 100 if total_cost > 0 else 0
            ws.cell(row=row, column=1, value=fert_type.replace('_', ' ').title()).border = border
            ws.cell(row=row, column=2, value=cost).border = border
            ws.cell(row=row, column=3, value=f"{percentage:.1f}%").border = border
            row += 1
        
        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    def _export_xml(self, data: Dict) -> str:
        """Export as XML"""
        
        root = ET.Element("FertilizerRecommendation")
        
        # Basic info
        basic_info = ET.SubElement(root, "BasicInformation")
        for key, value in data.get('basic_info', {}).items():
            elem = ET.SubElement(basic_info, key)
            elem.text = str(value)
        
        # Soil analysis
        soil_analysis = ET.SubElement(root, "SoilAnalysis")
        for key, value in data.get('soil_analysis', {}).items():
            elem = ET.SubElement(soil_analysis, key)
            elem.text = str(value)
        
        # Crop selection
        crop_selection = ET.SubElement(root, "CropSelection")
        for key, value in data.get('crop_selection', {}).items():
            elem = ET.SubElement(crop_selection, key)
            elem.text = str(value)
        
        # Nutrient requirements
        nutrients = ET.SubElement(root, "NutrientRequirements")
        for key, value in data.get('nutrient_requirements', {}).items():
            if isinstance(value, dict):
                sub_elem = ET.SubElement(nutrients, key)
                for sub_key, sub_value in value.items():
                    sub_sub_elem = ET.SubElement(sub_elem, sub_key)
                    sub_sub_elem.text = str(sub_value)
            else:
                elem = ET.SubElement(nutrients, key)
                elem.text = str(value)
        
        # Fertilizers
        fertilizers = ET.SubElement(root, "RecommendedFertilizers")
        for fert in data.get('recommended_fertilizers', []):
            fert_elem = ET.SubElement(fertilizers, "Fertilizer")
            for key, value in fert.items():
                elem = ET.SubElement(fert_elem, key)
                elem.text = str(value)
        
        # Application schedule
        schedule = ET.SubElement(root, "ApplicationSchedule")
        for app in data.get('application_schedule', []):
            app_elem = ET.SubElement(schedule, "Application")
            for key, value in app.items():
                elem = ET.SubElement(app_elem, key)
                elem.text = str(value)
        
        # Cost analysis
        cost = ET.SubElement(root, "CostAnalysis")
        cost_data = data.get('cost_analysis', {})
        for key, value in cost_data.items():
            if isinstance(value, dict):
                sub_elem = ET.SubElement(cost, key)
                for sub_key, sub_value in value.items():
                    sub_sub_elem = ET.SubElement(sub_elem, sub_key)
                    sub_sub_elem.text = str(sub_value)
            else:
                elem = ET.SubElement(cost, key)
                elem.text = str(value)
        
        # Format XML
        self._indent_xml(root)
        return ET.tostring(root, encoding='unicode')
    
    def _export_txt(self, data: Dict) -> str:
        """Export as text file"""
        
        output = []
        output.append("SMART FERTILIZER RECOMMENDATION REPORT")
        output.append("=" * 50)
        output.append("")
        
        # Basic information
        output.append("BASIC INFORMATION")
        output.append("-" * 20)
        basic_info = data.get('basic_info', {})
        for key, value in basic_info.items():
            output.append(f"{key.replace('_', ' ').title()}: {value}")
        output.append("")
        
        # Soil analysis
        output.append("SOIL ANALYSIS")
        output.append("-" * 15)
        soil_analysis = data.get('soil_analysis', {})
        for key, value in soil_analysis.items():
            output.append(f"{key.replace('_', ' ').title()}: {value}")
        output.append("")
        
        # Nutrient requirements
        output.append("NUTRIENT REQUIREMENTS")
        output.append("-" * 22)
        nutrients = data.get('nutrient_requirements', {})
        output.append(f"Nitrogen: {nutrients.get('total_nitrogen_kg_per_ha', 0)} kg/ha")
        output.append(f"Phosphorus: {nutrients.get('total_phosphorus_kg_per_ha', 0)} kg/ha")
        output.append(f"Potassium: {nutrients.get('total_potassium_kg_per_ha', 0)} kg/ha")
        output.append("")
        
        # Recommended fertilizers
        output.append("RECOMMENDED FERTILIZERS")
        output.append("-" * 25)
        fertilizers = data.get('recommended_fertilizers', [])
        for fert in fertilizers:
            output.append(f"• {fert.get('name', '')}")
            output.append(f"  N-P-K: {fert.get('n_content', 0)}-{fert.get('p_content', 0)}-{fert.get('k_content', 0)}")
            output.append(f"  Price: {fert.get('price_per_kg', 0)} per kg")
            output.append(f"  Availability: {fert.get('availability', '')}")
            output.append("")
        
        # Application schedule
        output.append("APPLICATION SCHEDULE")
        output.append("-" * 20)
        schedule = data.get('application_schedule', [])
        for app in schedule:
            output.append(f"• {app.get('stage', '').replace('_', ' ').title()}")
            output.append(f"  Days after planting: {app.get('days_after_planting', 0)}")
            output.append(f"  Fertilizer: {app.get('fertilizer_type', '')}")
            output.append(f"  Rate: {app.get('amount_kg_per_ha', 0)} kg/ha")
            output.append(f"  Method: {app.get('application_method', '').replace('_', ' ').title()}")
            if app.get('notes'):
                output.append(f"  Notes: {app.get('notes', '')}")
            output.append("")
        
        # Cost analysis
        output.append("COST ANALYSIS")
        output.append("-" * 15)
        cost_analysis = data.get('cost_analysis', {})
        output.append(f"Total Cost: {cost_analysis.get('currency', 'USD')} {cost_analysis.get('total_cost', 0)}")
        output.append(f"Cost per Hectare: {cost_analysis.get('currency', 'USD')} {cost_analysis.get('cost_per_hectare', 0)}")
        output.append("")
        
        return "\n".join(output)
    
    def _get_parameter_status(self, param: str, value: float) -> str:
        """Get parameter status classification"""
        
        if param == 'ph':
            if value < 5.5:
                return "Acidic"
            elif value > 7.5:
                return "Alkaline"
            else:
                return "Optimal"
        elif param == 'organic_matter':
            if value < 2.0:
                return "Low"
            elif value > 4.0:
                return "High"
            else:
                return "Medium"
        elif param in ['nitrogen', 'phosphorus', 'potassium']:
            thresholds = {
                'nitrogen': {'low': 200, 'high': 400},
                'phosphorus': {'low': 10, 'high': 50},
                'potassium': {'low': 100, 'high': 400}
            }
            thresh = thresholds.get(param, {'low': 0, 'high': 1000})
            if value < thresh['low']:
                return "Low"
            elif value > thresh['high']:
                return "High"
            else:
                return "Medium"
        elif param == 'cec':
            if value < 10:
                return "Low"
            elif value > 25:
                return "High"
            else:
                return "Medium"
        else:
            return "Normal"
    
    def _indent_xml(self, elem, level=0):
        """Indent XML for pretty printing"""
        i = "\n" + level * "  "
        if len(elem):
            if not elem.text or not elem.text.strip():
                elem.text = i + "  "
            if not elem.tail or not elem.tail.strip():
                elem.tail = i
            for child in elem:
                self._indent_xml(child, level + 1)
            if not child.tail or not child.tail.strip():
                child.tail = i
        else:
            if level and (not elem.tail or not elem.tail.strip()):
                elem.tail = i
    
    def create_export_package(self, recommendations: List[Dict], 
                            formats: List[str] = None) -> bytes:
        """Create a zip package with multiple export formats"""
        
        if formats is None:
            formats = ['json', 'csv', 'xlsx', 'txt']
        
        # Create zip buffer
        zip_buffer = io.BytesIO()
        
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for i, recommendation in enumerate(recommendations):
                rec_id = recommendation.get('recommendation_id', f'recommendation_{i+1}')
                
                for format_type in formats:
                    try:
                        if format_type == 'pdf':
                            # PDF requires bytes
                            content = self.export_recommendation(recommendation, format_type)
                            filename = f"{rec_id}.{format_type}"
                            zip_file.writestr(filename, content)
                        else:
                            # Text-based formats
                            content = self.export_recommendation(recommendation, format_type)
                            filename = f"{rec_id}.{format_type}"
                            if isinstance(content, bytes):
                                zip_file.writestr(filename, content)
                            else:
                                zip_file.writestr(filename, content.encode('utf-8'))
                    except Exception as e:
                        # Create error file if export fails
                        error_content = f"Error exporting {format_type}: {str(e)}"
                        zip_file.writestr(f"{rec_id}_{format_type}_error.txt", error_content)
        
        zip_buffer.seek(0)
        return zip_buffer.getvalue()
